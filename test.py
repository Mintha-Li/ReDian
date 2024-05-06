import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from systemfunc import create_table_header
from extract_top_rankings import extract_top_rankings

def create_table_header(platforms, num_ranks, output_file):
    # 创建一个工作簿
    wb = Workbook()

    # 选择默认的活动表单
    ws = wb.active

    # 表头信息
    header = ['平台']
    for platform in platforms:
        header.extend([platform] * num_ranks)

    # 添加表头和子表头
    for idx, value in enumerate(header, start=1):
        ws.cell(row=1, column=idx, value=value)

    # 在第二行第一列写入 '排名'
    ws.cell(row=2, column=1, value='排名')

    # 添加排名
    for col_idx in range(2, ws.max_column+1, (num_ranks)):
        for i in range(1, num_ranks+1):
            ws.cell(row=2, column=col_idx+i-1, value=i)

    # 合并单元格
    for col_idx in range(2, ws.max_column+1, (num_ranks)):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+num_ranks-1)

    # 保存工作簿
    wb.save(output_file)

def append_data_to_excel(input_file, source_column, rank_column, rank_num=3, output_file="platforms.xlsx"):
    # 读取需要排列的数据
    top_combined = extract_top_rankings(input_file, source_column, rank_column, rank_num)

    # 打开现有的 Excel 文件，并将新的数据追加到工作表的末尾
    with pd.ExcelWriter(output_file, mode='a', engine='openpyxl') as writer:
        writer.book = load_workbook(output_file)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        top_combined.to_excel(writer, index=False, header=False, startrow=2)

# 测试代码
input_file = "input.xlsx"  # 替换为你的输入文件路径
source_column = "source"
rank_column = "hot_num"
append_data_to_excel(input_file, source_column, rank_column)

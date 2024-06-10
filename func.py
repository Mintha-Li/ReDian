from openpyxl import Workbook
import os
from datetime import datetime
import pandas as pd


def chinese_weekday_date_display(date_obj):
    # 映射英文星期缩写到中文
    weekday_map = {
        'Mon': '一',
        'Tue': '二',
        'Wed': '三',
        'Thu': '四',
        'Fri': '五',
        'Sat': '六',
        'Sun': '日'
    }

    # 获取星期缩写并转换为中文
    weekday_abbr = date_obj.strftime('%a')
    chinese_weekday = weekday_map.get(weekday_abbr)

    # 构建日期显示字符串
    date_display = f"周{chinese_weekday}\n{date_obj.month}.{date_obj.day}"

    return date_display


def convert_to_wan_format(hot_num, source):
    """
    将热度值格式化为“xx万”。

    参数:
    hot_num (str/int/float): 热度值。
    source (str): 数据来源平台。

    返回:
    str: 格式化后的热度值。
    """
    if source == '知乎':
        # 移除“ 万热度”并保留“xx万”部分
        if ' 万热度' in hot_num:
            return hot_num.replace(' 万热度', '万')
    elif isinstance(hot_num, (int, float)):
        # 将微博和抖音的热度值转换为“xx万”格式
        return f"{hot_num / 10000:.0f}万" if hot_num >= 10000 else f"{hot_num}"
    return hot_num


def extract_top_rankings(input_file, source_column, rank_column, rank_num=3, platforms=['微博', '知乎', '抖音']):
    """
    提取各平台的前rank_num条数据,并将热度值格式化为“xx万”。

    参数:
    input_file (str): 输入的Excel文件路径。
    source_column (str): 平台列的列名。
    rank_column (str): 排名列的列名。
    rank_num (int): 每个平台提取的前几名数据。
    platforms (list): 要处理的平台列表。

    返回:
    pd.DataFrame: 格式化后的数据框。
    """

    # 读取Excel文件
    df = pd.read_excel(input_file)

    # 初始化空的数据框，用于存储结果
    top = pd.DataFrame()

    # 遍历每个平台，提取前rank_num条数据
    for platform in platforms:
        platform_data = df[df[source_column] == platform].nsmallest(rank_num, rank_column)
        top = pd.concat([top, platform_data])

    # 重置索引
    top_combined = top.reset_index(drop=True)

    # 根据平台对热度值进行格式化
    top_combined['hot_num'] = top_combined.apply(lambda row: convert_to_wan_format(row['hot_num'], row[source_column]),
                                                 axis=1)

    return top_combined


def process_files_in_folder(folder_path, output_folder, platforms=['微博', '知乎', '抖音'], num_ranks=3,
                            source_column='source', rank_column='rank'):
    """
    处理目标文件夹中的所有Excel文件,提取指定平台的排名数据并保存到新的Excel文件中。

    参数:
    folder_path (str): 目标文件夹路径,包含待处理的Excel文件。
    output_folder (str): 输出文件夹路径,用于保存处理后的Excel文件。
    platforms (list): 要处理的平台列表,默认包含['微博', '知乎', '抖音']。
    num_ranks (int): 每个平台提取的前几名数据,默认是3。
    source_column (str): 平台列的列名,默认是'source'。
    rank_column (str): 排名列的列名,默认是'rank'。
    """
    # 获取目标文件夹中的所有文件名，并按名称排序
    file_names = sorted(os.listdir(folder_path))

    # 确保文件夹中有文件
    if not file_names:
        print("文件夹中没有文件。")
        return

    # 创建新的Excel工作簿和工作表
    wb = Workbook()
    ws = wb.active

    # 表头信息初始化
    header = ['平台']
    for platform in platforms:
        header.extend([platform] * num_ranks)

    # 添加表头信息
    for idx, value in enumerate(header, start=2):
        ws.cell(row=1, column=idx, value=value)

    # 在第二行第一列写入'排名''
    ws.cell(row=2, column=2, value='排名')

    # 添加排名信息
    for col_idx in range(3, ws.max_column + 1, num_ranks):
        for i in range(1, num_ranks + 1):
            ws.cell(row=2, column=col_idx + i - 1, value=i)

    # 合并单元格，设置平台名称
    for col_idx in range(3, ws.max_column + 1, num_ranks):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + num_ranks - 1)

    # 遍历文件夹中的每个文件
    for file_index, file_name in enumerate(file_names):
        # 构建文件的完整路径
        file_path = os.path.join(folder_path, file_name)

        # 检查文件是否为 Excel 文件
        if file_name.endswith('.xlsx'):
            # 提取每个平台的排名数据
            top_combined = extract_top_rankings(file_path, source_column, rank_column, num_ranks, platforms=platforms)
            print(top_combined)

            date_str = file_name.split('-')[0]  # 假设文件名格式为 '2024060318-描述'
            date_obj = datetime.strptime(date_str, "%Y%m%d%H")
            date_display = f"{chinese_weekday_date_display(date_obj)}"

            #   计算当前行索引
            current_row = 3 + file_index * 3

            # 写入时间信息并合并单元格
            ws.cell(row=current_row, column=1, value=date_display)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=1)

            # 写入热搜标题
            ws.cell(row=current_row, column=2, value='标题')
            for col_index, title in enumerate(top_combined['title'], start=3):
                ws.cell(row=current_row, column=col_index, value=title)

            # 写入热度
            ws.cell(row=current_row + 1, column=2, value='热度')
            for col_index, hot_num in enumerate(top_combined['hot_num'], start=3):
                ws.cell(row=current_row + 1, column=col_index, value=hot_num)

            # 写入类型（空）
            ws.cell(row=current_row + 2, column=2, value='类型')
            for col_index in range(3, ws.max_column + 1):
                ws.cell(row=current_row + 2, column=col_index, value='')

    # 获取当前时间并格式化为字符串，作为文件名的一部分
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 获取文件夹中的第一个文件名和最后一个文件名（去掉扩展名）
    first_file_name = os.path.splitext(file_names[0].split('-')[0])[0]
    last_file_name = os.path.splitext(file_names[-1].split('-')[0])[0]

    # 构建输出文件路径
    output_file_name = f"{current_time}_{first_file_name}_to_{last_file_name}_热点.xlsx"
    output_file_path = os.path.join(output_folder, output_file_name)

    # 保存工作簿到输出文件路径
    wb.save(output_file_path)
    return output_file_name


if __name__ == '__main__':
    process_files_in_folder(folder_path='./input',
                            output_folder='./output',
                            platforms=['微博', '知乎', '抖音'],
                            num_ranks=3)

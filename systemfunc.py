import os
from datetime import datetime
from extract_top_ranking import extract_top_rankings
from openpyxl import Workbook


def process_files_in_folder(folder_path, output_folder,platforms = ['微博', '知乎', '抖音'],num_ranks = 3,source_column = 'source',rank_column = 'rank'):
    """
    处理目标文件夹中的所有Excel文件,提取指定平台的排名数据并保存到新的Excel文件中。

    参数:
    folder_path (str): 目标文件夹路径,包含待处理的Excel文件。
    output_folder (str): 输出文件夹路径,用于保存处理后的Excel文件。
    platforms (list): 要处理的平台列表,默认包含['微博', '知乎', '抖音']。
    num_ranks (int): 每个平台提取的前几名数据,默认是3。
    source_column (str): 平台列的列名,默认是'source'。
    rank_column (str): 排名列的列名,默认是'rank'。
    """
    # 获取目标文件夹中的所有文件名
    file_names = os.listdir(folder_path)

    # 创建新的Excel工作簿和工作表
    wb = Workbook()
    ws = wb.active

    # 表头信息初始化
    header = ['平台']
    for platform in platforms:
        header.extend([platform] * num_ranks)

    # 添加表头信息
    for idx, value in enumerate(header, start=1):
        ws.cell(row=1, column=idx, value=value)

    # 在第二行第一列写入'排名''
    ws.cell(row=2, column=1, value='排名')

    # 添加排名信息
    for col_idx in range(2, ws.max_column+1, (num_ranks)):
        for i in range(1, num_ranks+1):
            ws.cell(row=2, column=col_idx+i-1, value=i)

    # 合并单元格，设置平台名称
    for col_idx in range(2, ws.max_column+1, (num_ranks)):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+num_ranks-1)



    # 遍历文件夹中的每个文件
    for file_name in file_names:
        # 构建文件的完整路径
        file_path = os.path.join(folder_path, file_name)
        
        # 检查文件是否为 Excel 文件
        if file_name.endswith('.xlsx'):
            # 提取每个平台的排名数据
            top_combined = extract_top_rankings(file_path, source_column, rank_column,num_ranks,platforms=platforms)
            print(top_combined)
           
            #写入热搜标题
            ws.append(['标题'] + list(top_combined['title']))

            # 写入热度
            ws.append(['热度'] + list(top_combined['hot_num']))

            # 写入类型
            ws.append(['类型'])

    # 获取当前时间并格式化为字符串，作为文件名的一部分
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # 获取文件夹中的第一个文件名和最后一个文件名（去掉扩展名）
    first_file_name = os.path.splitext(file_names[0])[0]
    last_file_name = os.path.splitext(file_names[-1])[0]

    # 构建输出文件路径
    output_file_name = f"{current_time}_{first_file_name}_to_{last_file_name}_热点.xlsx"
    output_file_path = os.path.join(output_folder, output_file_name)

    # 保存工作簿到输出文件路径
    wb.save(output_file_path)